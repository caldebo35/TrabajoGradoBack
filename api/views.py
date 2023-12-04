from django.http import JsonResponse, HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action  # Agrega esta línea para importar action
from .models import *
from .serializer import *
from rest_framework.decorators import api_view, action
from rest_framework.response import Response
from rest_framework import status
import json
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
import os
from django.conf import settings
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import io
from django.http import FileResponse


@api_view(["GET"])
def apiOverview(request):
    api_urls = {
        "Categoria": "/categoria/",
        "Vehiculo": "/vehiculo/",
        "Bodega": "/bodega/",
        "Objeto": "/objeto/",
        "Objetos por categoria": "/categoria/<str:nombre_categoria>/",
        "Calculo": "/calculo/",
        "Generar Excel": "/generar_excel/",
        "Descargar Excel": "/descargar_excel/",
    }
    return Response(api_urls)


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer

    @action(detail=True, methods=["get"])
    def objetos_por_categoria(
        self, request, nombre_categoria=None
    ):  # Recibe el parámetro nombre_categoria
        try:
            categoria = Categoria.objects.get(
                nombre=nombre_categoria
            )  # Obtén la categoría por su nombre
            objetos = Objeto.objects.filter(categoria=categoria)
            serializer = ObjetoSerializer(objetos, many=True)
            return Response(serializer.data)
        except Categoria.DoesNotExist:
            return Response(
                {"message": "Categoría no encontrada"}, status=status.HTTP_404_NOT_FOUND
            )


@csrf_exempt
@api_view(["POST"])
def generar_excel(request):
    libro_excel = openpyxl.Workbook()
    hoja_excel = libro_excel.active
    if request.method == "POST":
        try:
            # Obtener datos del cuerpo de la solicitud
            data = json.loads(request.body.decode("utf-8"))

            # Verificar que "data" está presente y es una lista
            if "data" not in data or not isinstance(data["data"], list):
                return JsonResponse(
                    {"error": "Formato de datos incorrecto"}, status=400
                )

            # Sección: Cantidad total de artículos y Volumen en m3
            cantidad_total = sum(item.get("cantidad", 0) for item in data["data"])
            volumen_total = sum(item.get("volumen", 0) for item in data["data"])
            nombre_vehiculo = data['data'][2]['vehiculo']['nombre']
            capacidad_min = data['data'][2]['vehiculo']['capacidad_min']
            capacidad_max = data['data'][2]['vehiculo']['capacidad_max']
            nombre_bodega = data['data'][3]['bodega']['nombre']
            area_bodega = data['data'][3]['bodega']['area']
            volumen_bodega = data['data'][3]['bodega']['volumen']
            objetos_seleccionados = data['data'][4].get('objetos', [])
            
            # Ajustar el ancho de la columna A a 5.14
            hoja_excel.column_dimensions['A'].width = 5.14
            hoja_excel.column_dimensions['B'].width = 13.71
            hoja_excel.column_dimensions['C'].width = 25.29
            hoja_excel.column_dimensions['D'].width = 5.29
            hoja_excel.column_dimensions['E'].width = 34.86
            hoja_excel.column_dimensions['F'].width = 14.29

            

            # Agregar encabezados y valores
            hoja_excel['B1'] = 'Mi inventario en TrasteaT'
            hoja_excel['B3'] = 'Cantidad total de artículos'
            hoja_excel['D3'] = cantidad_total
            hoja_excel['E3'] = 'Volumen total'
            hoja_excel['F3'] = volumen_total
            hoja_excel['G3'] = 'm3'

            hoja_excel['B5'] = 'Vehículo Requerido'
            hoja_excel['D5'] = nombre_vehiculo
            hoja_excel['F5'] = capacidad_min
            hoja_excel['G5'] = '-'
            hoja_excel['H5'] = capacidad_max
            hoja_excel['I5'] = 'T'
            
            hoja_excel['B6'] = 'Bodega Requerida'
            hoja_excel['D6'] = nombre_bodega
            hoja_excel['F6'] = 'Area'
            hoja_excel['G6'] = area_bodega
            hoja_excel['H6'] = 'Volumen'
            hoja_excel['I6'] = volumen_bodega
            hoja_excel['J6'] = 'm3'

            hoja_excel['B7'] = 'Inventario'
            hoja_excel['B8'] = 'Cantidad'
            hoja_excel['D8'] = 'Objetos'

            # Agregar datos del inventario a la hoja
            for i, objeto in enumerate(objetos_seleccionados, start=10):
                nombre_objeto = objeto.get('nombre', '')
                cantidad_objeto = objeto.get('cantidad', 0)

                hoja_excel[f'D{i}'] = nombre_objeto
                hoja_excel[f'B{i}'] = cantidad_objeto


            # Combinar celdas para el título
            hoja_excel.merge_cells('B1:G1')
            hoja_excel.merge_cells('B7:G7')
            hoja_excel.merge_cells('B3:C3')
            hoja_excel.merge_cells('B5:C5')
            hoja_excel.merge_cells('D5:E5')
            hoja_excel.merge_cells('B6:C6')
            hoja_excel.merge_cells('D6:E6')
            hoja_excel.merge_cells('B8:C8')
            hoja_excel.merge_cells('D8:E8')
            hoja_excel.merge_cells('F8:G8')
            hoja_excel['B1'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['B7'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['B3'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['B5'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['B6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['B8'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['D8'].alignment = Alignment(horizontal='center', vertical='center')
            for row_number in range(8, 110):
                hoja_excel.merge_cells(f'B{row_number}:C{row_number}')
                hoja_excel.merge_cells(f'D{row_number}:E{row_number}')
                hoja_excel.merge_cells(f'F{row_number}:G{row_number}')
                
                hoja_excel[f'B{row_number}'].alignment = Alignment(horizontal='center', vertical='center')
                hoja_excel[f'D{row_number}'].alignment = Alignment(horizontal='center', vertical='center')
                hoja_excel[f'F{row_number}'].alignment = Alignment(horizontal='center', vertical='center')

            hoja_excel['B1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['B3'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['C3'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['D3'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['E3'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['F3'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['G3'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['H3'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['I3'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['B6'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['C6'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['D6'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['E6'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['F6'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['G6'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['H6'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['I6'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['B5'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['C5'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['D5'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['E5'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['F5'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['G5'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['H5'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['I5'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['B7'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['B8'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['C8'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['D8'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['E8'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['F8'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            hoja_excel['G8'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            # Establecer el tamaño y el tipo de letra para toda la hoja
            color_letra = 'F9C307'
            for row in hoja_excel.iter_rows(min_row=1, max_col=hoja_excel.max_column, max_row=hoja_excel.max_row):
                for cell in row:
                    cell.font = Font(size=14, name='Comic Sans MS', color=color_letra)

            # for objeto in data['data'][4]['objetos']:
            #     nombre_objeto = objeto['nombre']
            #     cantidad_objeto = objeto['cantidad']
            #     volumen_objeto = objeto['volumen']
            # print("Nombre del objeto:", nombre_objeto)
            # print("Cantidad del objeto:", cantidad_objeto)
            # print("Volumen del objeto:", volumen_objeto)
            

            excel_buffer = io.BytesIO()
            libro_excel.save(excel_buffer)

            # Devolver el archivo Excel como respuesta
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f"attachment; filename=inventario.xlsx"

            # Guardar el archivo en el sistema de archivos si es necesario
            excel_file_path = os.path.join(settings.MEDIA_ROOT, f"inventario.xlsx")
            with open(excel_file_path, "wb") as file:
                file.write(excel_buffer.getvalue())

            # Devolver una respuesta exitosa
            return response

        except Exception as e:
            # Manejar errores
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)


def descargar_excel(request):
    excel_file_path = os.path.join(settings.MEDIA_ROOT, "inventario.xlsx")

    if os.path.exists(excel_file_path):
        with open(excel_file_path, "rb") as file:
            response = HttpResponse(
                file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=inventario.xlsx"

        return response
    else:
        return JsonResponse({"error": "El archivo Excel no existe"}, status=404)


class VehiculoViewSet(viewsets.ModelViewSet):
    queryset = Vehiculo.objects.all()
    serializer_class = VehiculoSerializer


class BodegaViewSet(viewsets.ModelViewSet):
    queryset = Bodega.objects.all()
    serializer_class = BodegaSerializer


class ObjetoViewSet(viewsets.ModelViewSet):
    queryset = Objeto.objects.all()
    serializer_class = ObjetoSerializer


class CalculoViewSet(viewsets.ModelViewSet):
    queryset = Calculo.objects.all()
    serializer_class = CalculoSerializer
